from fastapi import APIRouter, UploadFile, File
import docx
import openpyxl
import io
import uuid
import os
from app.services.ai_provider import AIProvider
from app.services.import_pipeline import refine_schema_with_ai

router = APIRouter()

@router.post("/docx")
async def parse_docx(file: UploadFile = File(...)):
    content = await file.read()
    doc = docx.Document(io.BytesIO(content))
    
    raw_fields = []
    warnings = []
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
            
        # Deterministic parsing: heading styles become sections
        is_heading = para.style.name.startswith('Heading')
        
        raw_fields.append({
            "id": str(uuid.uuid4()),
            "key": f"field_{len(raw_fields)}",
            "label": text[:100], 
            "raw_text": text,
            "is_section": is_heading,
            "required": '*' in text
        })

    # Hybrid AI Inference
    ai_provider = AIProvider()
    model = os.getenv("AI_MODEL", "gpt-4o")
    
    refined_fields = await refine_schema_with_ai(raw_fields, ai_provider, model)
    
    # Clean up non-schema fields
    for f in refined_fields:
        f.pop('raw_text', None)
        f.pop('is_section', None)

    return {
        "schema": {
            "version": "1.0.0",
            "metadata": {"title": file.filename},
            "fields": refined_fields,
            "layout": {"sections": [{"id": "s1", "fields": [f["id"] for f in refined_fields if f.get("type") != "section_heading"]}]}
        },
        "warnings": warnings
    }

@router.post("/xlsx")
async def parse_xlsx(file: UploadFile = File(...)):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active
    
    raw_fields = []
    warnings = []
    headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []
    
    # Peek at row 2 for type inference
    sample_data = [cell.value for cell in sheet[2]] if sheet.max_row > 1 else []
    
    for idx, header in enumerate(headers):
        if not header: continue
        
        sample_val = str(sample_data[idx]) if idx < len(sample_data) and sample_data[idx] is not None else ""
        
        raw_fields.append({
            "id": str(uuid.uuid4()),
            "key": f"col_{idx}",
            "label": str(header),
            "raw_text": f"Sample data: {sample_val}",
            "is_section": False,
            "required": '*' in str(header)
        })

    # Hybrid AI Inference
    ai_provider = AIProvider()
    model = os.getenv("AI_MODEL", "gpt-4o")
    
    refined_fields = await refine_schema_with_ai(raw_fields, ai_provider, model)
    
    # Clean up non-schema fields
    for f in refined_fields:
        f.pop('raw_text', None)
        f.pop('is_section', None)

    return {
        "schema": {
            "version": "1.0.0",
            "metadata": {"title": file.filename},
            "fields": refined_fields,
            "layout": {"sections": [{"id": "s1", "fields": [f["id"] for f in refined_fields]}]}
        },
        "warnings": warnings
    }
